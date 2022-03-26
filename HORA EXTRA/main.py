import pandas as pd
import numpy as np
from datetime import timedelta, datetime
import holidays
import math
from openpyxl import Workbook, load_workbook
from  openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import requests
from bs4 import BeautifulSoup



ORANGE = Color(rgb='FFf79646')

def get_salario_min():
    response = requests.get('https://www.google.com/search?client=firefox-b-d&q=salario+minimo+2022')
    html_byte = response.content
    i = str(html_byte).find('R$')
    salario_min = str(html_byte)[i:i+10].replace(' ', '').replace('R$', '').replace('.', '').replace(',', '.')
    for character in salario_min:
        if not character.isdigit() and character != ',' and character != '.':
            salario_min = salario_min.replace(character, '')
    return float(salario_min)

def adic_noturno(entrada, saida):

    ad_noturno_saida=22*60*60
    ad_noturno_entrada=5*60*60
    ratio_saida = saida//(29*60*60)
    if saida>ad_noturno_saida:
        if ratio_saida != 0:
            h_noturno_saida = saida-ad_noturno_saida-(saida-ad_noturno_saida-(7*60*60)*ratio_saida)
        else:
            h_noturno_saida = saida-ad_noturno_saida
    else:
        h_noturno_saida = 0
    if entrada<ad_noturno_entrada:
        h_noturno_entrada = ad_noturno_entrada-entrada
    else:
        h_noturno_entrada = 0
    if saida<ad_noturno_entrada:
        h_noturno_saida=saida-entrada
        h_noturno_entrada = 0
    h_noturno = h_noturno_saida+h_noturno_entrada
    return h_noturno

def horas_50_100(entrada, saida, obs, jornada_de_trabalho):
    if obs == 'DOMINGO' or obs == 'FERIADO':
        h_100 = saida-entrada
        h_50 = 0
    elif obs=='SABADO':
        h_100 = 0
        h_50 = saida-entrada
        if jornada_de_trabalho == 32400: # trabalha 4 horas no sabado
            h_50 = saida-entrada-(4*60*60)
    else:
        h_100=0
        h_50 = saida-entrada-jornada_de_trabalho
    return h_50, h_100

def get_obs(dia, uf):
    lista_feriados = list(feriados.loc[feriados.loc[:, 'UF']==uf, 'DATA'].apply(lambda x: datetime.timestamp(x)).values)
    dia_timestamp = datetime.timestamp(dia)
    if dia_timestamp in lista_feriados:
        obs = 'FERIADO'
        print('FERIADO', feriados.loc[np.logical_and(feriados.loc[:, 'UF']==uf, feriados.loc[:, 'DATA']==dia), 'FERIADO'])
    elif dia.weekday() == 5:
        obs = 'SABADO'
    elif dia.weekday() == 6:
        obs = 'DOMINGO'
    else:
        obs = None
    return obs

def format_time(seconds):
    if seconds >= 0:
        conversion = timedelta(seconds=seconds)
        f_hours = str(conversion)
        if len(f_hours) == 7:
            f_hours = f'0{f_hours}'
    elif seconds < 0:
        conversion = timedelta(seconds=-seconds)
        f_hours = str(conversion)
        if len(f_hours) == 7:
            f_hours = f'-0{f_hours}'
        else:
            f_hours = f'-{f_hours}'
    if f_hours[0].isdigit() and 'day' in f_hours:
        f_hours_splitado = f_hours.split('day')
        dias = int(f_hours_splitado[0].replace(' ', ''))
        tempo = f_hours.split('day')[1].replace(' ', '').replace(',', '')
        print(tempo)
        if 's' in tempo:
            tempo = tempo[1:]
        if ':' in tempo[:2]:
            t = tempo[:1]
        else:
            t = tempo[:2]
        tempo1 = int(t)
        for i in range(0, dias):
            tempo1+=24
        if tempo[2] == ':':
            f_hours = f'{tempo1}{tempo[2:]}'
        else:
            f_hours = f'{tempo1}:{tempo[2:]}'

    return f_hours

def alimentar_resumo(nome, salario, ad_not_formated, h_50_formated, h_100_formated, h_noturno, h_50, h_100, resumo_detalhado, valor_h):

    valor_noturno = valor_h*0.2*h_noturno
    valor_50 = valor_h*1.5*h_50
    valor_100 = valor_h*2*h_100
    dct = {
    'NOME':nome,
    'VALOR HORA': valor_h,
    'HORAS NOTURNO': ad_not_formated,
    'VALOR NOTURNO': valor_noturno,
    'HORAS 50%': h_50_formated,
    'VALOR 50%': valor_50,
    'HORAS 100%': h_100_formated,
    'VALOR 100%': valor_100,
    'TOTAL':valor_noturno+valor_50+valor_100,
    'RESUMO DETALHADO':resumo_detalhado,
    }
    global resumo
    resumo = resumo.append([dct,], ignore_index=True, sort=False)
    return

def format_cell(c):
    c.font = Font(b=True)
    c.alignment = Alignment(vertical='center', horizontal='center')
    c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
    return

def gerar_especie(wb, resumo, troco):

    ws = wb.create_sheet('ESPECIE')
    thick = Side(style='thick')
    thin = Side(style='thin')
    ############################################################################
    # HEADER
    ws.merge_cells('A1:E1')
    ws['A1'] = 'ESPÉCIE'
    cells_header_title = [ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1']]
    for c in cells_header_title:                  #  00000000
        c.font = Font(b=True)
        c.alignment = Alignment(vertical='center', horizontal='center')
        c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
        c.border = Border(top=thick, bottom=thick)
    ws['A1'].border = Border(left=thick, top=thick, bottom=thick)
    ws['E1'].border = Border(right=thick, top=thick, bottom=thick)
    ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'] = 'NOME', 'VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL'
    cells_header = [ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2']]
    for c in cells_header:                  #  00000000
        c.font = Font(b=True)
        c.alignment = Alignment(vertical='center', horizontal='center')
        c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ############################################################################
    # DATA
    global especie
    especie = resumo.loc[:, ['NOME','VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']].copy()
    especie.loc[:, ['VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']] = especie.loc[:, ['VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']].apply(lambda x: np.ceil(x/2))
    for row in especie.iterrows():
        for col in ['VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']:
            while especie.loc[row[0], col]%troco !=0:
                especie.loc[row[0], col] += 0.01
                especie.loc[row[0], col] = np.ceil(especie.loc[row[0], col])

        print(especie)
        i=row[0]+3
        ws[f'A{i}'] = row[1]['NOME']
        ws[f'B{i}'] = row[1]['VALOR NOTURNO']
        ws[f'B{i}'].number_format = '"R$"#,##0.00'
        ws[f'C{i}'] = row[1]['VALOR 50%']
        ws[f'C{i}'].number_format = '"R$"#,##0.00'
        ws[f'D{i}'] = row[1]['VALOR 100%']
        ws[f'D{i}'].number_format = '"R$"#,##0.00'
        ws[f'E{i}'] = row[1]['TOTAL']
        ws[f'E{i}'].number_format = '"R$"#,##0.00'
        cells_of_the_row = [ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'], ws[f'E{i}']]
        for c in cells_of_the_row:
            c.font = Font(b=False)
            c.alignment = Alignment(vertical='center', horizontal='center')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    global folha
    folha = resumo
    folha.loc[:, ['VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']] -= especie.loc[:, ['VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL']]
    ############################################################################
    # FOOTER
    i = i+1
    ws[f'A{i}'] = 'TOTAL'
    ws[f'B{i}'] = f'=SUM(B3:B{i-1})'
    ws[f'B{i}'].number_format = '"R$"#,##0.00'
    ws[f'C{i}'] = f'=SUM(C3:C{i-1})'
    ws[f'C{i}'].number_format = '"R$"#,##0.00'
    ws[f'D{i}'] = f'=SUM(D3:D{i-1})'
    ws[f'D{i}'].number_format = '"R$"#,##0.00'
    ws[f'E{i}'] = f'=SUM(E3:E{i-1})'
    ws[f'E{i}'].number_format = '"R$"#,##0.00'
    cells_footer_total = [ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'], ws[f'E{i}']]
    for c in cells_footer_total:                  #  00000000
        format_cell(c)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)

    ############################################################################
    # AUTOADJUST WIDTH COLUMNS
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.remove(wb['Sheet'])
    wb.save(f'./RESUMOS/RESUMO.xlsx')
    return

def gerar_resumo_detalhado(nome, salario, resumo_detalhado, totais, valor_h, risco, salario_min, jornada_de_trabalho):
    wb = Workbook()
    ws = wb.create_sheet(nome)
    thick = Side(style='thick')
    thin = Side(style='thin')
    ws.merge_cells('A1:J1')
    ws['A1'] = nome
    cells_header_title = [ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'],  ws['I1'], ws['J1']]
    for c in cells_header_title:                  #  00000000
            c.font = Font(b=True)
            c.alignment = Alignment(vertical='center', horizontal='center')
            c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
            c.border = Border(top=thick, bottom=thick)
    ws['A1'].border = Border(left=thick, top=thick, bottom=thick)
    ws['J1'].border = Border(right=thick, top=thick, bottom=thick)

    ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'], ws['G2'], ws['H2'], ws['I2'], ws['J2'] = 'SALÁRIO:', salario, 'RISCO:', risco, 'SALARIO MINIMO ATUAL:', salario_min, 'VALOR HORA:', valor_h, 'JORNADA DE TRABALHO:', format_time(jornada_de_trabalho)
    lst_left = [ws['A2'], ws['C2'], ws['E2'], ws['G2'], ws['I2']]
    for c in lst_left:                  #  00000000
            c.font = Font(b=True)
            c.alignment = Alignment(vertical='center', horizontal='right')
            c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
            c.border = Border(left=thick, top=thick, bottom=thick)
    lst_right = [ws['B2'], ws['D2'], ws['F2'], ws['H2'], ws['J2']]
    for c in lst_right:                  #  00000000
            c.font = Font(b=True)
            c.alignment = Alignment(vertical='center', horizontal='left')
            c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
            c.border = Border(right=thick, top=thick, bottom=thick)
    lst_monetary = [ws['B2'], ws['F2'], ws['H2']]
    for c in lst_monetary:
        c.number_format = '"R$"#,##0.00'
    ws['A3'], ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'], ws['G3'], ws['H3'], ws['I3'], ws['J3'] = 'DATA', 'FERIADO', 'ENTRADA', 'SAÍDA', 'ADICIONAL NOTURNO', 'VALOR NOTURNO', 'HORAS 50%', 'VALOR 50%', 'HORAS 100%', 'VALOR 100%'

    cells_header = [ws['A3'], ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'], ws['G3'], ws['H3'], ws['I3'], ws['J3']]
    for c in cells_header:                  #  00000000
            c.font = Font(b=True)
            c.alignment = Alignment(vertical='center', horizontal='center')
            c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
            c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ############################################################################
    # DATA
    for row in resumo_detalhado.copy().reset_index().iterrows():
        i=row[0]+4
        ws[f'A{i}'] = row[1]['DATA']
        ws[f'B{i}'] = row[1]['FERIADO']
        ws[f'C{i}'] = row[1]['ENTRADA']
        ws[f'D{i}'] = row[1]['SAÍDA']
        ws[f'E{i}'] = row[1]['ADICIONAL NOTURNO']
        ws[f'F{i}'] = row[1]['ADNOT']*0.2*valor_h
        ws[f'F{i}'].number_format = '"R$"#,##0.00'
        ws[f'G{i}'] = row[1]['HORAS 50%']
        ws[f'H{i}'] = row[1]['H50']*1.5*valor_h
        ws[f'H{i}'].number_format = '"R$"#,##0.00'
        ws[f'I{i}'] = row[1]['HORAS 100%']
        ws[f'J{i}'] = row[1]['H100']*2*valor_h
        ws[f'J{i}'].number_format = '"R$"#,##0.00'
        cells_of_the_row = [ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'], ws[f'E{i}'], ws[f'F{i}'], ws[f'G{i}'], ws[f'H{i}'], ws[f'I{i}'], ws[f'J{i}']]
        for c in cells_of_the_row:
            c.font = Font(b=False)
            c.alignment = Alignment(vertical='center', horizontal='center')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ############################################################################
    # FOOTER
    i = i+1
    ws.merge_cells(f'A{i}:D{i}')
    ws[f'A{i}'] = 'TOTAL'
    ws[f'E{i}'] = format_time(totais['ADNOT']*3600)
    ws[f'F{i}'] = f'=SUM(F4:F{i-1})'
    ws[f'F{i}'].number_format = '"R$"#,##0.00'
    ws[f'G{i}'] = format_time(totais['H50']*3600)
    ws[f'H{i}'] = f'=SUM(H4:H{i-1})'
    ws[f'H{i}'].number_format = '"R$"#,##0.00'
    ws[f'I{i}'] = format_time(totais['H100']*3600)
    ws[f'J{i}'] = f'=SUM(J4:J{i-1})'
    ws[f'J{i}'].number_format = '"R$"#,##0.00'
    cells_footer_total_title = [ws[f'B{i}'], ws[f'C{i}']]
    for c in cells_footer_total_title:
        format_cell(c)
        c.border = Border(top=thick, bottom=thick)
    format_cell(ws[f'A{i}'])
    ws[f'A{i}'].border = Border(left=thick, top=thick, bottom=thick)
    format_cell(ws[f'D{i}'])
    ws[f'D{i}'].border = Border(right=thick, top=thick, bottom=thick)
    cells_footer_total = [ws[f'E{i}'], ws[f'F{i}'], ws[f'G{i}'], ws[f'H{i}'], ws[f'I{i}'], ws[f'J{i}']]
    for c in cells_footer_total:                  #  00000000
        format_cell(c)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ############################################################################
    # AUTOADJUST WIDTH COLUMNS
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.remove(wb['Sheet'])
    wb.save(f'./RESUMOS/{nome} detalhado.xlsx')

def gerar_resumo(resumo):
    wb = Workbook()
    ws = wb.create_sheet('RESUMO')
    thick = Side(style='thick')
    thin = Side(style='thin')
    ############################################################################
    # HEADER
    ws.merge_cells('A1:I1')
    ws['A1'] = 'RESUMO'
    cells_header_title = [ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'], ws['I1']]
    for c in cells_header_title:                  #  00000000
        c.font = Font(b=True)
        c.alignment = Alignment(vertical='center', horizontal='center')
        c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
        c.border = Border(top=thick, bottom=thick)
    ws['A1'].border = Border(left=thick, top=thick, bottom=thick)
    ws['I1'].border = Border(right=thick, top=thick, bottom=thick)
    ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'], ws['G2'], ws['H2'], ws['I2'] = 'NOME', 'VALOR HORA', 'HORAS NOTURNO', 'VALOR NOTURNO', 'HORAS 50%', 'VALOR 50%', 'HORAS 100%', 'VALOR 100%', 'TOTAL'
    cells_header = [ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'], ws['G2'], ws['H2'], ws['I2']]
    for c in cells_header:                  #  00000000
        c.font = Font(b=True)
        c.alignment = Alignment(vertical='center', horizontal='center')
        c.fill = PatternFill(patternType='solid', fgColor=ORANGE)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ############################################################################
    # DATA
    for row in resumo.copy().reset_index().iterrows():
        i=row[0]+3
        ws[f'A{i}'] = row[1]['NOME']
        ws[f'B{i}'] = row[1]['VALOR HORA']
        ws[f'B{i}'].number_format = '"R$"#,##0.00'
        ws[f'C{i}'] = row[1]['HORAS NOTURNO']
        ws[f'D{i}'] = row[1]['VALOR NOTURNO']
        ws[f'D{i}'].number_format = '"R$"#,##0.00'
        ws[f'E{i}'] = row[1]['HORAS 50%']
        ws[f'F{i}'] = row[1]['VALOR 50%']
        ws[f'F{i}'].number_format = '"R$"#,##0.00'
        ws[f'G{i}'] = row[1]['HORAS 100%']
        ws[f'H{i}'] = row[1]['VALOR 100%']
        ws[f'H{i}'].number_format = '"R$"#,##0.00'
        ws[f'I{i}'] = row[1]['TOTAL']
        ws[f'I{i}'].number_format = '"R$"#,##0.00'
        cells_of_the_row = [ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'], ws[f'E{i}'], ws[f'F{i}'], ws[f'G{i}'], ws[f'H{i}'], ws[f'I{i}']]
        for c in cells_of_the_row:
            c.font = Font(b=False)
            c.alignment = Alignment(vertical='center', horizontal='center')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ############################################################################
    # FOOTER
    i = i+1
    ws.merge_cells(f'A{i}:C{i}')
    ws[f'A{i}'] = 'TOTAL'
    ws[f'D{i}'] = f'=SUM(D3:D{i-1})'
    ws[f'D{i}'].number_format = '"R$"#,##0.00'
    ws[f'F{i}'] = f'=SUM(F3:F{i-1})'
    ws[f'F{i}'].number_format = '"R$"#,##0.00'
    ws[f'H{i}'] = f'=SUM(H3:H{i-1})'
    ws[f'H{i}'].number_format = '"R$"#,##0.00'
    ws[f'I{i}'] = f'=SUM(I3:I{i-1})'
    ws[f'I{i}'].number_format = '"R$"#,##0.00'

    format_cell(ws[f'B{i}'])
    ws[f'B{i}'].border = Border(top=thick, bottom=thick)
    format_cell(ws[f'A{i}'])
    ws[f'A{i}'].border = Border(left=thick, top=thick, bottom=thick)
    format_cell(ws[f'D{i}'])
    ws[f'B{i}'].border = Border(right=thick, top=thick, bottom=thick)
    cells_footer_total = [ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'], ws[f'E{i}'], ws[f'F{i}'], ws[f'G{i}'], ws[f'H{i}'], ws[f'I{i}']]
    for c in cells_footer_total:                  #  00000000
        format_cell(c)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    gerar_especie(wb, resumo, 2)
    ############################################################################
    # AUTOADJUST WIDTH COLUMNS
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    ############################################################################

def get_salario_risco(nome):
    df = pd.read_excel('//solsrv1/dp$/FUNCIONARIOS/FUNCIONARIOS.xlsx')
    func = df.loc[df.loc[:, 'Unnamed: 2'] == nome, :]
    if len(func) != 1:
        print(func)
        print('FUNCIOANRIOS ENCONTRADOS DIFERENTE DE 1!!!!!')
    salario = func['Unnamed: 25'].values[0]
    risco = func['Unnamed: 28'].values[0]
    entrada =  func['Unnamed: 22'].values[0]
    if type(entrada) == float:
        print(entrada)
    entrada_sec = (entrada.hour*60*60)+(entrada.minute*60)
    saida = func['Unnamed: 23'].values[0]
    saida_sec = (saida.hour*60*60)+(saida.minute*60)
    jornada_de_trabalho = saida_sec-entrada_sec
    uf = func['Unnamed: 29'].values[0]
    return salario, risco, jornada_de_trabalho, uf

def gerar_relatorio_detalhado():
    df = pd.read_excel('./HE Model.xlsx')
    c = len(df.columns)-1
    global resumo
    resumo = pd.DataFrame(columns=['NOME', 'VALOR HORA','HORAS NOTURNO', 'VALOR NOTURNO', 'HORAS 50%', 'VALOR 50%', 'HORAS 100%', 'VALOR 100%'])
    global especie
    especie = pd.DataFrame(columns=['NOME','VALOR NOTURNO', 'VALOR 50%', 'VALOR 100%', 'TOTAL'])
    global feriados
    feriados = pd.read_excel('./HE Model.xlsx', sheet_name='FERIADOS')


    for loop in range(-c, 0, 2):
        if loop != -2:
            data = df.iloc[:,loop:loop+2].copy()
        else:
            data = df.iloc[:, loop:].copy()

        nome = data.columns[0]
        print(nome)
        data.dropna(how='all', axis=0, inplace=True)
        if not data.empty:
            data.columns = ['ENTRADA', 'SAÍDA']
            data = data.drop(0, axis=0)
            dias = df.loc[:, 'Unnamed: 0'].drop(0, axis=0)
            dias.name='DATA'
            data[['ADICIONAL NOTURNO', 'HORAS 50%', 'HORAS 100%', 'DATA']] = [np.nan, np.nan, np.nan, np.nan]
            salario, risco, jornada_de_trabalho, uf = get_salario_risco(nome)
            for row in data.iterrows():
                dia = dias.loc[row[0]]
                entrada = timedelta(hours=data.loc[row[0], 'ENTRADA'].hour, minutes=data.loc[row[0], 'ENTRADA'].minute).seconds
                saida = timedelta(hours=data.loc[row[0], 'SAÍDA'].hour, minutes=data.loc[row[0], 'SAÍDA'].minute).seconds
                if saida<entrada:
                    print('saida ->', saida)
                    saida+=24*60*60
                obs = get_obs(dia, uf)
                if obs == 'FERIADO':
                    feriado = holidays.BR(state='SP').get(dia)
                else:
                    feriado = 'Nenhum'
                h_noturno = adic_noturno(entrada, saida)
                h_50, h_100  = horas_50_100(entrada, saida, obs, jornada_de_trabalho)
                ad_not_formated, h_50_formated, h_100_formated = format_time(h_noturno), format_time(h_50), format_time(h_100)
                data.loc[row[0], ['ADICIONAL NOTURNO', 'HORAS 50%', 'HORAS 100%', 'ADNOT', 'H50', 'H100', 'DATA', 'FERIADO']] = [ad_not_formated, h_50_formated, h_100_formated, h_noturno/3600, h_50/3600, h_100/3600, dia.strftime('%d/%m/%Y'), feriado]
            totais = data.loc[:, ['ADNOT', 'H50', 'H100']].sum(numeric_only=True)
            resumo_detalhado = data.copy()
            ad_not_formated, h_50_formated, h_100_formated = format_time(totais['ADNOT']*3600), format_time(totais['H50']*3600), format_time(totais['H100']*3600)

            salario_min= get_salario_min()
            if risco == 'NENHUM':
                valor_h = salario/220
            elif risco == 'INSALUBRIDADE':
                valor_h = ((salario_min*0.2)+salario)/220
            elif risco == 'PERICULOSIDADE':
                valor_h = (salario*1.3)/220

            alimentar_resumo(nome, salario, ad_not_formated, h_50_formated, h_100_formated, totais['ADNOT'], totais['H50'], totais['H100'], resumo_detalhado, valor_h)
            gerar_resumo_detalhado(nome, salario, resumo_detalhado, totais, valor_h, risco, salario_min, jornada_de_trabalho)
    gerar_resumo(resumo)
    return

def _validate_nomes():
    df_ver = pd.read_excel('./HE Model.xlsx', sheet_name='RASCUNHO 2')
    df = pd.read_excel('//solsrv1/dp$/FUNCIONARIOS/FUNCIONARIOS.xlsx')
    func = df.loc[:, 'Unnamed: 2']
    status = []
    nomes = []
    f = func.unique()
    for nome in df_ver['NOME'].unique():
        nomes.append(nome)
        if nome in f:
            status.append(True)
        else:
            status.append(False)

    df_0 = pd.DataFrame({'NOMES':nomes, 'STATUS':status,})
    df_0.to_excel('./VALIDATE.xlsx')
################################################################################
# GERAR RELATÓRIO
# gerar_relatorio_detalhado()
